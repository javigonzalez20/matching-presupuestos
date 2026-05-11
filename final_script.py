#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""

Prototipo final TFM:
- Lee un Excel de presupuesto nuevo
- Lee uno o varios Excels históricos
- Extrae filas reales de conceptos
- Recupera Top-5 con multilingual-e5-base
- Reordena con cross-encoder ms-marco-MiniLM-L-12-v2
- Aplica regla final de medida:
    * dentro del ranking del cross-encoder,
      elige el primer candidato con medida compatible
    * si no existe ninguno, elige el primero del ranking cross
- Genera una hoja AI por cada hoja de entrada tipo ELEC/MEC

Uso:
python final_script.py ^
  --input "nuevo_proyecto.xlsx" ^
  --historical "historicos\\*.xlsx" ^
  --output "salida_ai.xlsx"

Dependencias:
pip install pandas openpyxl numpy scikit-learn unidecode sentence-transformers torch
"""

from __future__ import annotations

import argparse
import glob
import logging
import os
import re
import shutil
from dataclasses import dataclass
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from unidecode import unidecode

HAS_ST = False
SentenceTransformer = None
CrossEncoder = None

try:
    from sentence_transformers import SentenceTransformer, CrossEncoder
    HAS_ST = True
except Exception:
    HAS_ST = False


logging.basicConfig(
    level=logging.INFO,
    format="[%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)


@dataclass
class BudgetItem:
    source_file: str
    sheet_name: str
    row_index_excel: int

    ref: str
    unidad: str
    concepto: str
    cantidad: Optional[float]

    interno: str = ""
    descripcion_larga: str = ""

    mo_unidad: Optional[float] = None
    mat_euros: Optional[float] = None
    mat_mxn: Optional[float] = None
    mat_mex: Optional[float] = None
    mat_dop: Optional[float] = None
    mat_usd: Optional[float] = None
    mat_unidad: Optional[float] = None
    mat_trans_adu: Optional[float] = None
    pvp_unid: Optional[float] = None
    pvp_total_euros: Optional[float] = None

    concepto_norm: str = ""
    medidas: Tuple[str, ...] = ()

    def full_text(self) -> str:
        parts = [self.concepto]
        if self.descripcion_larga:
            parts.append(self.descripcion_larga)
        if self.interno:
            parts.append(self.interno)
        return " | ".join([p for p in parts if p])


@dataclass
class MatchCandidate:
    item: BudgetItem
    emb_score: float
    cross_score: Optional[float] = None
    rank_cross: Optional[int] = None
    medida_coincide: bool = False


@dataclass
class FinalMatch:
    query_item: BudgetItem
    matched_item: Optional[BudgetItem]
    selected_by_measure: bool
    criterio: str
    emb_top1_score: Optional[float]
    final_cross_score: Optional[float]
    debug_top_candidates: List[str]


def normalize_text(text: object) -> str:
    if text is None:
        return ""
    text = str(text).strip()
    if not text:
        return ""

    text = unidecode(text.lower())

    text = text.replace("ø", " diametro ")
    text = text.replace("º", " ")
    text = text.replace('"', " pulgadas ")
    text = text.replace("”", " pulgadas ")
    text = text.replace("“", " ")
    text = text.replace("´", " ")
    text = text.replace("`", " ")
    text = text.replace("_", " ")
    text = text.replace("×", " x ")
    text = text.replace("–", "-")
    text = text.replace("—", "-")
    text = text.replace("\\", " ")
    text = text.replace("/", " / ")

    text = re.sub(r"[^a-z0-9\.\,\-\+\sx/]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    return text


def parse_number(value: object) -> Optional[float]:
    if value is None:
        return None

    if isinstance(value, (int, float, np.integer, np.floating)):
        if pd.isna(value):
            return None
        return float(value)

    s = str(value).strip()
    if not s:
        return None

    s = s.replace("$", "").replace("€", "").replace("%", "")
    s = s.replace("MXN", "").replace("USD", "").replace("DOP", "")
    s = s.replace("EUR", "").replace("Ud", "").replace("ud", "")
    s = s.strip()

    s = re.sub(r"[^0-9,\.\-]", "", s)

    if not s:
        return None

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")

    try:
        return float(s)
    except ValueError:
        return None


def clean_measure_token(token: str) -> str:
    token = unidecode(token.lower().strip())
    token = token.replace(" ", "")
    token = token.replace(",", ".")
    return token


def extract_measures(text: str) -> Tuple[str, ...]:
    if not text:
        return tuple()

    t = normalize_text(text)
    found: List[str] = []

    patterns = [
        r"\b\d+\s*/\s*\d+\b",
        r"\b\d+x\d+\s+\d+/\d+\b",
        r"\b\d+x\d+(?:[.,]\d+)?\b",
        r"\b\d+(?:[.,]\d+)?\s*mm\b",
        r"\bawg\s*[-]?\s*\d+(?:/\d+)?\b",
        r"\b\d+(?:[.,]\d+)?\b",
    ]

    for pat in patterns:
        for m in re.finditer(pat, t):
            token = clean_measure_token(m.group(0))
            if token and token not in found:
                found.append(token)

    return tuple(found)

HEADER_SYNONYMS = {
    "ref": ["ref", "referencia"],
    "unidad": ["unidad", "und", "u", "ud"],
    "concepto": ["concepto", "descripcion", "descripción", "item", "nombre"],
    "cantidad": ["cantidad", "cant"],
    "interno": ["interno", "codigo interno", "código interno", "codigo", "código"],
    "mo_unidad": ["m.o. unidad", "mo unidad", "mano de obra unidad"],
    "mat_euros": ["mat. euros", "mat euros"],
    "mat_mxn": ["mat. mxn", "mat mxn"],
    "mat_mex": ["mat. $ mex", "mat $ mex", "mat mex"],
    "mat_dop": ["mat dop", "mat. dop"],
    "mat_usd": ["mat. $", "mat $", "mat usd"],
    "mat_unidad": ["mat. unidad", "mat unidad"],
    "mat_trans_adu": ["mat. + trans y adu", "mat + trans y adu"],
    "pvp_unid": ["pvp unid", "pvp unid.", "pvp unidad"],
    "pvp_total_euros": ["pvp total euros", "pvp total"],
    "coment": ["coment", "comentario", "comentarios"],
}


def normalize_header_name(value: object) -> str:
    return normalize_text(value)


def score_header_row(row_values: Sequence[object]) -> int:
    norm = [normalize_header_name(v) for v in row_values]
    hits = 0
    must_have = 0

    for cell in norm:
        if "concepto" in cell:
            hits += 5
            must_have += 1
        if cell == "ref":
            hits += 4
        if "unidad" in cell:
            hits += 3
        if "cantidad" in cell:
            hits += 3
        if "interno" in cell:
            hits += 2
        if "pvp" in cell:
            hits += 1
        if "mat" in cell:
            hits += 1
        if "m.o" in cell or cell == "mo":
            hits += 1

    if must_have == 0:
        return -1
    return hits


def find_header_row(df_raw: pd.DataFrame, max_scan_rows: int = 40) -> int:
    best_idx = -1
    best_score = -999

    for idx in range(min(max_scan_rows, len(df_raw))):
        score = score_header_row(df_raw.iloc[idx].tolist())
        if score > best_score:
            best_score = score
            best_idx = idx

    if best_idx < 0:
        raise ValueError("No se pudo detectar automáticamente la fila de cabecera.")

    logger.info(f"Fila de cabecera detectada en índice {best_idx} (0-based).")
    return best_idx


def map_columns(headers: List[str]) -> Dict[str, Optional[int]]:
    mapping: Dict[str, Optional[int]] = {k: None for k in HEADER_SYNONYMS.keys()}

    for idx, header in enumerate(headers):
        h = normalize_header_name(header)
        for target, synonyms in HEADER_SYNONYMS.items():
            for syn in synonyms:
                if normalize_text(syn) == h or normalize_text(syn) in h:
                    if mapping[target] is None:
                        mapping[target] = idx

    return mapping


def get_cell(row: Sequence[object], idx: Optional[int]) -> object:
    if idx is None:
        return None
    if idx < 0 or idx >= len(row):
        return None
    return row[idx]

def is_section_header(ref: str, concepto: str, cantidad: Optional[float]) -> bool:
    ref = (ref or "").strip()
    concepto = (concepto or "").strip()

    if not concepto:
        return True

    if re.fullmatch(r"\d{2}(?:\.\d{2})?", ref) and len(concepto) < 80:
        return True

    if cantidad is None and len(concepto) < 80 and ref and not re.search(r"[A-Za-z]", ref):
        return True

    return False


def is_long_description_row(ref: str, concepto: str, cantidad: Optional[float]) -> bool:
    ref = (ref or "").strip()
    concepto = (concepto or "").strip()

    if not concepto:
        return False

    prefixes = [
        "SUMINISTRO",
        "INSTALACION",
        "INSTALACIÓN",
        "MONTAJE",
        "CIRCUITO ELECTRICO",
        "PUNTOS DE ALIMENTACIÓN",
        "BANDEJA REPARTIDORA",
        "PATCH CORDS",
        "TRANSCEIVER",
        "COMPUESTO POR LOS SIGUIENTES ELEMENTOS",
    ]

    if any(concepto.upper().startswith(p) for p in prefixes):
        if not ref or len(ref) < 3:
            return True

    if len(concepto) > 90 and cantidad is None and not ref:
        return True

    return False


def is_real_concept_row(ref: str, concepto: str, cantidad: Optional[float]) -> bool:
    ref = (ref or "").strip()
    concepto = (concepto or "").strip()

    if not concepto:
        return False
    if is_section_header(ref, concepto, cantidad):
        return False
    if is_long_description_row(ref, concepto, cantidad):
        return False

    if ref and len(concepto) >= 3:
        return True
    if cantidad is not None and len(concepto) >= 3:
        return True

    return False


def parse_budget_sheet(excel_path: str, sheet_name: Optional[str] = None) -> List[BudgetItem]:
    xls = pd.ExcelFile(excel_path)
    target_sheet = sheet_name or xls.sheet_names[0]

    df_raw = pd.read_excel(excel_path, sheet_name=target_sheet, header=None, dtype=object)

    header_row_idx = find_header_row(df_raw)
    headers = [str(v).strip() if v is not None else "" for v in df_raw.iloc[header_row_idx].tolist()]
    mapping = map_columns(headers)

    logger.info(f"Column mapping para '{os.path.basename(excel_path)}' / '{target_sheet}': {mapping}")

    items: List[BudgetItem] = []
    last_main_idx: Optional[int] = None

    for i in range(header_row_idx + 1, len(df_raw)):
        row = df_raw.iloc[i].tolist()

        ref = str(get_cell(row, mapping["ref"]) or "").strip()
        unidad = str(get_cell(row, mapping["unidad"]) or "").strip()
        concepto = str(get_cell(row, mapping["concepto"]) or "").strip()
        cantidad = parse_number(get_cell(row, mapping["cantidad"]))
        interno = str(get_cell(row, mapping["interno"]) or "").strip()

        mo_unidad = parse_number(get_cell(row, mapping["mo_unidad"]))
        mat_euros = parse_number(get_cell(row, mapping["mat_euros"]))
        mat_mxn = parse_number(get_cell(row, mapping["mat_mxn"]))
        mat_mex = parse_number(get_cell(row, mapping["mat_mex"]))
        mat_dop = parse_number(get_cell(row, mapping["mat_dop"]))
        mat_usd = parse_number(get_cell(row, mapping["mat_usd"]))
        mat_unidad = parse_number(get_cell(row, mapping["mat_unidad"]))
        mat_trans_adu = parse_number(get_cell(row, mapping["mat_trans_adu"]))
        pvp_unid = parse_number(get_cell(row, mapping["pvp_unid"]))
        pvp_total_euros = parse_number(get_cell(row, mapping["pvp_total_euros"]))

        if is_real_concept_row(ref, concepto, cantidad):
            item = BudgetItem(
                source_file=os.path.basename(excel_path),
                sheet_name=target_sheet,
                row_index_excel=i + 1,
                ref=ref,
                unidad=unidad,
                concepto=concepto,
                cantidad=cantidad,
                interno=interno,
                mo_unidad=mo_unidad,
                mat_euros=mat_euros,
                mat_mxn=mat_mxn,
                mat_mex=mat_mex,
                mat_dop=mat_dop,
                mat_usd=mat_usd,
                mat_unidad=mat_unidad,
                mat_trans_adu=mat_trans_adu,
                pvp_unid=pvp_unid,
                pvp_total_euros=pvp_total_euros,
            )
            item.concepto_norm = normalize_text(item.full_text())
            item.medidas = extract_measures(item.concepto_norm)
            items.append(item)
            last_main_idx = len(items) - 1

        elif is_long_description_row(ref, concepto, cantidad) and last_main_idx is not None:
            desc = concepto.strip()
            if desc:
                prev = items[last_main_idx]
                prev.descripcion_larga += (" " if prev.descripcion_larga else "") + desc
                prev.concepto_norm = normalize_text(prev.full_text())
                prev.medidas = extract_measures(prev.concepto_norm)

    logger.info(
        f"Extraídos {len(items)} conceptos útiles de '{os.path.basename(excel_path)}' / '{target_sheet}'."
    )
    return items



class MatchingEngine:
    def __init__(
        self,
        embedding_model_name: str = "intfloat/multilingual-e5-base",
        cross_encoder_name: str = "cross-encoder/ms-marco-MiniLM-L-12-v2",
        top_k: int = 5,
    ) -> None:
        self.embedding_model_name = embedding_model_name
        self.cross_encoder_name = cross_encoder_name
        self.top_k = top_k

        self.use_sentence_transformers = HAS_ST
        self.embedding_model = None
        self.cross_encoder = None

        self.historical_items: List[BudgetItem] = []
        self.historical_matrix = None

        self.tfidf_vectorizer = None
        self.tfidf_hist = None

    def load_models(self) -> None:
        if not self.use_sentence_transformers:
            logger.warning("sentence-transformers no disponible. Fallback a TF-IDF.")
            return

        try:
            logger.info(f"Cargando embedding model: {self.embedding_model_name}")
            self.embedding_model = SentenceTransformer(self.embedding_model_name)
        except Exception as e:
            logger.warning(f"No se pudo cargar el embedding model. Fallback a TF-IDF. Error: {e}")
            self.use_sentence_transformers = False
            self.embedding_model = None

        if self.use_sentence_transformers:
            try:
                logger.info(f"Cargando cross-encoder: {self.cross_encoder_name}")
                self.cross_encoder = CrossEncoder(self.cross_encoder_name)
            except Exception as e:
                logger.warning(f"No se pudo cargar el cross-encoder. Se omitirá reranking. Error: {e}")
                self.cross_encoder = None

    def fit_historical(self, historical_items: List[BudgetItem]) -> None:
        self.historical_items = historical_items
        corpus = [it.concepto_norm for it in historical_items]

        if self.use_sentence_transformers and self.embedding_model is not None:
            logger.info(f"Calculando embeddings históricos para {len(corpus)} conceptos...")
            self.historical_matrix = self.embedding_model.encode(
                corpus,
                show_progress_bar=True,
                convert_to_numpy=True,
                normalize_embeddings=True,
            )
        else:
            logger.info("Construyendo TF-IDF histórico...")
            self.tfidf_vectorizer = TfidfVectorizer(ngram_range=(1, 2))
            self.tfidf_hist = self.tfidf_vectorizer.fit_transform(corpus)

    def _embedding_candidates(self, query_text: str) -> List[MatchCandidate]:
        if self.use_sentence_transformers and self.embedding_model is not None:
            q_emb = self.embedding_model.encode(
                [query_text],
                convert_to_numpy=True,
                normalize_embeddings=True,
            )[0]
            sims = np.dot(self.historical_matrix, q_emb)
        else:
            q_vec = self.tfidf_vectorizer.transform([query_text])
            sims = cosine_similarity(q_vec, self.tfidf_hist)[0]

        top_idx = np.argsort(-sims)[: self.top_k]

        candidates = [
            MatchCandidate(
                item=self.historical_items[idx],
                emb_score=float(sims[idx])
            )
            for idx in top_idx
        ]
        return candidates

    def _rerank(self, query_item: BudgetItem, candidates: List[MatchCandidate]) -> List[MatchCandidate]:
        if not candidates or self.cross_encoder is None:
            for rank, cand in enumerate(candidates, start=1):
                cand.rank_cross = rank
            return candidates

        pairs = [(query_item.concepto_norm, cand.item.concepto_norm) for cand in candidates]
        scores = self.cross_encoder.predict(pairs)

        for cand, score in zip(candidates, scores):
            cand.cross_score = float(score)

        candidates = sorted(
            candidates,
            key=lambda x: x.cross_score if x.cross_score is not None else -999999,
            reverse=True,
        )

        for rank, cand in enumerate(candidates, start=1):
            cand.rank_cross = rank

        return candidates

    @staticmethod
    def _measure_match_exact_logic(
        query_measures: Tuple[str, ...],
        cand_measures: Tuple[str, ...],
    ) -> bool:
        if not query_measures or not cand_measures:
            return False

        q = set(m.strip() for m in query_measures if m and str(m).strip())
        c = set(m.strip() for m in cand_measures if m and str(m).strip())

        return len(q.intersection(c)) > 0

    def select_final(self, query_item: BudgetItem, candidates: List[MatchCandidate]) -> FinalMatch:
        if not candidates:
            return FinalMatch(
                query_item=query_item,
                matched_item=None,
                selected_by_measure=False,
                criterio="sin_candidatos",
                emb_top1_score=None,
                final_cross_score=None,
                debug_top_candidates=[],
            )

        for cand in candidates:
            cand.medida_coincide = self._measure_match_exact_logic(
                query_item.medidas,
                cand.item.medidas,
            )

        debug = []
        for cand in candidates:
            debug.append(
                f"rank_cross={cand.rank_cross} | match={cand.item.concepto} | "
                f"emb={cand.emb_score:.4f} | cross={cand.cross_score} | "
                f"medida_coincide={cand.medida_coincide}"
            )

        compatibles = [cand for cand in candidates if cand.medida_coincide]

        if len(compatibles) > 0:
            elegido = compatibles[0]
            criterio = "cross_con_medida"
            selected_by_measure = True
        else:
            elegido = candidates[0]
            criterio = "cross_sin_medida"
            selected_by_measure = False

        return FinalMatch(
            query_item=query_item,
            matched_item=elegido.item,
            selected_by_measure=selected_by_measure,
            criterio=criterio,
            emb_top1_score=candidates[0].emb_score if candidates else None,
            final_cross_score=elegido.cross_score,
            debug_top_candidates=debug,
        )

    def match_one(self, query_item: BudgetItem) -> FinalMatch:
        candidates = self._embedding_candidates(query_item.concepto_norm)
        candidates = self._rerank(query_item, candidates)
        return self.select_final(query_item, candidates)

    def match_many(self, query_items: List[BudgetItem]) -> List[FinalMatch]:
        results = []
        for i, q in enumerate(query_items, start=1):
            if i % 20 == 0:
                logger.info(f"Matching {i}/{len(query_items)}...")
            results.append(self.match_one(q))
        return results


def autosize_worksheet_columns(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)


def write_output_excel_sheet(workbook_path: str, output_sheet_name: str, matches: List[FinalMatch]) -> None:
    wb = load_workbook(workbook_path)

    if output_sheet_name in wb.sheetnames:
        del wb[output_sheet_name]

    ws = wb.create_sheet(output_sheet_name)

    headers = [
        "source_file",
        "sheet_name",
        "row_index_excel",
        "ref",
        "unidad",
        "cantidad",
        "concepto_nuevo",
        "descripcion_larga_nuevo",
        "interno_nuevo",
        "medidas_query",
        "criterio_final",
        "seleccionado_por_medida",
        "concepto_historico",
        "descripcion_larga_historico",
        "ref_historico",
        "unidad_historico",
        "interno_historico",
        "medidas_historico",
        "emb_top1_score",
        "cross_score_final",
        "mo_unidad_historico",
        "mat_euros_historico",
        "mat_mxn_historico",
        "mat_mex_historico",
        "mat_dop_historico",
        "mat_usd_historico",
        "mat_unidad_historico",
        "mat_trans_adu_historico",
        "pvp_unid_historico",
        "pvp_total_euros_historico",
        "debug_top_candidates",
    ]

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for r, fm in enumerate(matches, start=2):
        q = fm.query_item
        h = fm.matched_item

        row = {
            "source_file": q.source_file,
            "sheet_name": q.sheet_name,
            "row_index_excel": q.row_index_excel,
            "ref": q.ref,
            "unidad": q.unidad,
            "cantidad": q.cantidad,
            "concepto_nuevo": q.concepto,
            "descripcion_larga_nuevo": q.descripcion_larga,
            "interno_nuevo": q.interno,
            "medidas_query": ", ".join(q.medidas),
            "criterio_final": fm.criterio,
            "seleccionado_por_medida": fm.selected_by_measure,
            "concepto_historico": h.concepto if h else "",
            "descripcion_larga_historico": h.descripcion_larga if h else "",
            "ref_historico": h.ref if h else "",
            "unidad_historico": h.unidad if h else "",
            "interno_historico": h.interno if h else "",
            "medidas_historico": ", ".join(h.medidas) if h else "",
            "emb_top1_score": fm.emb_top1_score,
            "cross_score_final": fm.final_cross_score,
            "mo_unidad_historico": h.mo_unidad if h else None,
            "mat_euros_historico": h.mat_euros if h else None,
            "mat_mxn_historico": h.mat_mxn if h else None,
            "mat_mex_historico": h.mat_mex if h else None,
            "mat_dop_historico": h.mat_dop if h else None,
            "mat_usd_historico": h.mat_usd if h else None,
            "mat_unidad_historico": h.mat_unidad if h else None,
            "mat_trans_adu_historico": h.mat_trans_adu if h else None,
            "pvp_unid_historico": h.pvp_unid if h else None,
            "pvp_total_euros_historico": h.pvp_total_euros if h else None,
            "debug_top_candidates": " || ".join(fm.debug_top_candidates),
        }

        for c, hname in enumerate(headers, start=1):
            ws.cell(row=r, column=c, value=row[hname])

    autosize_worksheet_columns(ws)
    wb.save(workbook_path)
    logger.info(f"Hoja '{output_sheet_name}' guardada en: {workbook_path}")



def expand_historical_paths(paths: Sequence[str]) -> List[str]:
    files: List[str] = []
    for p in paths:
        expanded = glob.glob(p)
        if expanded:
            files.extend(expanded)
        else:
            files.append(p)

    files = [f for f in files if os.path.isfile(f)]
    files = list(dict.fromkeys(files))
    return files


def main() -> None:
    parser = argparse.ArgumentParser(description="Prototipo final TFM para matching de presupuestos Excel.")
    parser.add_argument("--input", required=True, help="Excel del nuevo proyecto.")
    parser.add_argument(
        "--historical",
        nargs="+",
        required=True,
        help="Uno o varios Excels históricos. Acepta comodines, por ejemplo historicos/*.xlsx"
    )
    parser.add_argument("--output", default="salida_ai.xlsx", help="Excel de salida.")
    parser.add_argument("--input-sheet", default=None, help="Nombre de hoja concreta del Excel de entrada.")
    parser.add_argument("--historical-sheet", default=None, help="Nombre de hoja concreta en históricos.")
    parser.add_argument("--top-k", type=int, default=5, help="Top-k inicial recuperado con embeddings.")
    parser.add_argument(
        "--embedding-model",
        default="intfloat/multilingual-e5-base",
        help="Modelo de embedding. Default = mejor modelo estudiado."
    )
    parser.add_argument(
        "--cross-encoder",
        default="cross-encoder/ms-marco-MiniLM-L-12-v2",
        help="Cross-encoder. Default = mejor cross-encoder estudiado."
    )

    args = parser.parse_args()

    if not os.path.isfile(args.input):
        raise FileNotFoundError(f"No existe el Excel de entrada: {args.input}")

    hist_files = expand_historical_paths(args.historical)
    if not hist_files:
        raise FileNotFoundError("No se encontraron Excels históricos.")

    logger.info("=== DETECTANDO HOJAS DEL PRESUPUESTO NUEVO ===")
    xls_input = pd.ExcelFile(args.input)

    if args.input_sheet:
        input_sheets_to_process = [args.input_sheet]
    else:
        input_sheets_to_process = [
            s for s in xls_input.sheet_names
            if ("ELEC" in s.upper() or "MEC" in s.upper())
        ]

    if not input_sheets_to_process:
        raise RuntimeError(
            "No se encontraron hojas de entrada tipo ELEC o MEC. "
            "Indica una hoja manualmente con --input-sheet."
        )

    logger.info(f"Hojas de entrada a procesar: {input_sheets_to_process}")

    logger.info("=== PARSEANDO HISTÓRICOS ===")
    historical_items: List[BudgetItem] = []
    for hf in hist_files:
        try:
            if args.historical_sheet:
                sheet_names = [args.historical_sheet]
            else:
                xls_hist = pd.ExcelFile(hf)
                sheet_names = xls_hist.sheet_names

            for sheet_name in sheet_names:
                try:
                    items = parse_budget_sheet(hf, sheet_name=sheet_name)
                    historical_items.extend(items)
                except Exception as e:
                    logger.warning(f"No se pudo procesar histórico '{hf}' hoja '{sheet_name}': {e}")

        except Exception as e:
            logger.warning(f"No se pudo abrir histórico '{hf}': {e}")

    if not historical_items:
        raise RuntimeError("No se extrajeron conceptos útiles de los históricos.")

    logger.info(f"Total conceptos históricos cargados: {len(historical_items)}")

    logger.info("=== INICIALIZANDO MOTOR DE MATCHING ===")
    engine = MatchingEngine(
        embedding_model_name=args.embedding_model,
        cross_encoder_name=args.cross_encoder,
        top_k=args.top_k,
    )
    engine.load_models()
    engine.fit_historical(historical_items)

    logger.info("=== PREPARANDO ARCHIVO DE SALIDA ===")
    if os.path.abspath(args.input) != os.path.abspath(args.output):
        shutil.copyfile(args.input, args.output)
    else:
        logger.warning(
            "El archivo de salida coincide con el de entrada. "
            "Se escribirán las hojas AI dentro del mismo workbook."
        )

    logger.info("=== EJECUTANDO PIPELINE FINAL POR HOJAS ===")
    processed_any = False

    for input_sheet in input_sheets_to_process:
        logger.info(f"--- Procesando hoja: {input_sheet} ---")

        query_items = parse_budget_sheet(args.input, sheet_name=input_sheet)

        if not query_items:
            logger.warning(f"No se extrajeron conceptos útiles de la hoja '{input_sheet}'. Se omite.")
            continue

        results = engine.match_many(query_items)

        output_sheet_name = f"AI_{input_sheet.replace(' ', '_')}"
        output_sheet_name = output_sheet_name[:31]

        write_output_excel_sheet(
            workbook_path=args.output,
            output_sheet_name=output_sheet_name,
            matches=results,
        )
        processed_any = True

    if not processed_any:
        raise RuntimeError("No se pudo procesar ninguna hoja de entrada útil.")

    logger.info("Proceso finalizado correctamente.")


if __name__ == "__main__":
    main()